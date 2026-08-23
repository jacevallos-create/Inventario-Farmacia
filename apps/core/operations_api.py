from io import BytesIO
from datetime import timedelta
from uuid import uuid4

from django.db import transaction
from django.db.models import DecimalField, ExpressionWrapper, F, Sum
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.auditoria.models import AuditLog
from apps.cajas.models import MovimientoCaja, SesionCaja
from apps.farmacias.models import Farmacia
from apps.inventario.models import InventarioFarmacia, MovimientoInventario
from apps.lotes.models import Lote
from apps.ventas.models import DevolucionVenta, DevolucionVentaLote, NotaCredito, Venta, VentaLote


def audit(request, action, entity, instance, description, pharmacy=None, changes=None):
    forwarded = request.META.get("HTTP_X_FORWARDED_FOR", "").split(",")[0].strip()
    AuditLog.objects.create(
        usuario=request.user, farmacia=pharmacy, accion=action, entidad=entity,
        objeto_id=str(instance.pk), descripcion=description, cambios=changes or {},
        ip=forwarded or request.META.get("REMOTE_ADDR"),
    )


class SaleCreateView(APIView):
    permission_classes = [IsAuthenticated]

    @transaction.atomic
    def post(self, request):
        branch_code = str(request.data.get("branchId", "")).strip()
        medicine_id = request.data.get("productId")
        try:
            quantity = int(request.data.get("qty", 0))
        except (TypeError, ValueError):
            quantity = 0
        if quantity < 1:
            return Response({"detail": "La cantidad debe ser mayor que cero."}, status=400)

        pharmacy = Farmacia.objects.filter(codigo__iexact=branch_code, activo=True).first()
        if not pharmacy:
            return Response({"detail": "La sucursal no existe o está inactiva."}, status=404)
        if not (request.user.is_staff or request.user.is_superuser or request.user.asignaciones_farmacia.filter(farmacia=pharmacy, activo=True).exists()):
            return Response({"detail": "No tienes acceso a esta sucursal."}, status=403)

        inventory = InventarioFarmacia.objects.select_for_update().select_related("medicamento").filter(
            farmacia=pharmacy, medicamento_id=medicine_id
        ).first()
        if not inventory or inventory.stock_actual < quantity:
            return Response({"detail": "No hay existencias suficientes."}, status=409)
        payment = str(request.data.get("payment") or "EFECTIVO").upper()
        if payment not in ("EFECTIVO", "TARJETA", "TRANSFERENCIA"):
            return Response({"detail": "La forma de pago no es válida."}, status=400)
        cash_session = SesionCaja.objects.select_for_update().filter(usuario=request.user, caja__farmacia=pharmacy, cerrada_en__isnull=True).first()
        if payment == "EFECTIVO" and not cash_session:
            return Response({"detail": "Debes abrir una caja antes de registrar ventas en efectivo."}, status=409)

        today = timezone.localdate()
        lots = list(Lote.objects.select_for_update().filter(
            farmacia=pharmacy, medicamento_id=medicine_id,
            cantidad_disponible__gt=0, fecha_vencimiento__gte=today,
        ).order_by("fecha_vencimiento", "id"))
        if sum(item.cantidad_disponible for item in lots) < quantity:
            return Response({"detail": "No hay lotes vigentes suficientes. Los lotes vencidos están bloqueados."}, status=409)

        reference = str(request.data.get("reference") or uuid4().hex)
        sale = Venta.objects.create(
            referencia_cliente=reference, farmacia=pharmacy, medicamento=inventory.medicamento,
            usuario=request.user, cliente_nombre=request.data.get("customer") or "Consumidor final",
            cantidad=quantity, precio_unitario=inventory.precio_venta,
            total=inventory.precio_venta * quantity,
            forma_pago=payment, sesion_caja=cash_session,
        )
        remaining = quantity
        previous_total = inventory.stock_actual
        consumed = []
        for lot in lots:
            if not remaining:
                break
            used = min(remaining, lot.cantidad_disponible)
            lot.cantidad_disponible -= used
            lot.save(update_fields=["cantidad_disponible", "actualizado_en"])
            VentaLote.objects.create(venta=sale, lote=lot, cantidad=used, precio_unitario=inventory.precio_venta)
            MovimientoInventario.objects.create(
                farmacia=pharmacy, medicamento=inventory.medicamento, lote=lot,
                tipo=MovimientoInventario.Tipo.SALIDA, concepto="VENTA", cantidad=used,
                saldo_anterior=previous_total, saldo_nuevo=previous_total - used,
                costo_unitario=lot.costo_unitario, documento_tipo="VENTA", documento_id=sale.pk,
                usuario=request.user, observacion=f"Venta {reference} por FEFO",
            )
            previous_total -= used
            remaining -= used
            consumed.append({"number": lot.numero, "quantity": used, "expires": lot.fecha_vencimiento.isoformat()})

        inventory.stock_actual -= quantity
        inventory.save(update_fields=["stock_actual", "actualizado_en"])
        if cash_session:
            MovimientoCaja.objects.create(sesion=cash_session, tipo=MovimientoCaja.Tipo.VENTA, forma_pago=payment, monto=sale.total, referencia=reference, usuario=request.user, observacion=f"Venta {reference}")
        audit(request, AuditLog.Accion.CREAR, "venta", sale, f"Venta {reference}", pharmacy, {"cantidad": quantity, "lotes": consumed})
        return Response({
            "sale": {"id": sale.referencia_cliente, "branchId": pharmacy.codigo.lower(),
                     "product": inventory.medicamento.nombre_comercial, "sku": inventory.medicamento.codigo_interno,
                     "qty": quantity, "total": float(sale.total), "customer": sale.cliente_nombre,
                     "date": timezone.localtime(sale.creado_en).strftime("%d/%m/%Y, %H:%M"), "lots": consumed,
                     "user": request.user.email, "payment": payment, "cancelled": False},
            "stock": inventory.stock_actual,
        }, status=201)


class SaleCancelView(APIView):
    permission_classes = [IsAuthenticated]
    @transaction.atomic
    def post(self, request, sale_id):
        if not (request.user.is_staff or request.user.is_superuser):
            return Response({"detail": "La anulación requiere autorización administrativa."}, status=403)
        sale = Venta.objects.select_for_update().select_related("farmacia", "medicamento", "sesion_caja").filter(referencia_cliente=sale_id).first()
        reason = str(request.data.get("reason") or "").strip()
        if not sale or sale.anulada or not reason:
            return Response({"detail": "Venta no disponible o motivo obligatorio ausente."}, status=409)
        inventory = InventarioFarmacia.objects.select_for_update().get(farmacia=sale.farmacia, medicamento=sale.medicamento)
        previous = inventory.stock_actual
        for detail in sale.lotes_consumidos.select_related("lote"):
            lot = Lote.objects.select_for_update().get(pk=detail.lote_id)
            lot.cantidad_disponible += detail.cantidad; lot.save()
            MovimientoInventario.objects.create(farmacia=sale.farmacia, medicamento=sale.medicamento, lote=lot, tipo="ENTRADA", concepto="ANULACION_VENTA", cantidad=detail.cantidad, saldo_anterior=inventory.stock_actual, saldo_nuevo=inventory.stock_actual + detail.cantidad, costo_unitario=lot.costo_unitario, documento_tipo="VENTA", documento_id=sale.id, usuario=request.user, observacion=reason)
            inventory.stock_actual += detail.cantidad
        inventory.save()
        if sale.sesion_caja and not sale.sesion_caja.cerrada_en:
            MovimientoCaja.objects.create(sesion=sale.sesion_caja, tipo=MovimientoCaja.Tipo.DEVOLUCION, forma_pago=sale.forma_pago, monto=sale.total, referencia=sale.referencia_cliente, usuario=request.user, observacion=f"Anulación: {reason}")
        sale.anulada=True; sale.anulada_por=request.user; sale.anulada_en=timezone.now(); sale.motivo_anulacion=reason; sale.save()
        audit(request, AuditLog.Accion.MODIFICAR, "venta", sale, f"Venta {sale.referencia_cliente} anulada", sale.farmacia, {"motivo": reason, "stock_anterior": previous, "stock_nuevo": inventory.stock_actual})
        return Response({"id": sale.referencia_cliente, "cancelled": True, "stock": inventory.stock_actual})


class CustomerReturnView(APIView):
    permission_classes = [IsAuthenticated]
    @transaction.atomic
    def post(self, request, sale_id):
        if not (request.user.is_staff or request.user.is_superuser):
            return Response({"detail": "La devolución requiere autorización administrativa."}, status=403)
        sale = Venta.objects.select_for_update().select_related("farmacia", "medicamento", "sesion_caja").filter(referencia_cliente=sale_id, anulada=False).first()
        reason = str(request.data.get("reason") or "").strip()
        try: quantity = int(request.data.get("quantity", 0))
        except (TypeError, ValueError): quantity = 0
        returned = sum(x.cantidad for x in sale.devoluciones.all()) if sale else 0
        if not sale or not reason or quantity < 1 or quantity > sale.cantidad - returned:
            return Response({"detail": "Cantidad no disponible o motivo obligatorio ausente."}, status=409)
        refund = sale.precio_unitario * quantity
        returned_record = DevolucionVenta.objects.create(venta=sale, usuario=request.user, cantidad=quantity, motivo=reason, autorizada_por=request.user, total_devuelto=refund)
        inventory = InventarioFarmacia.objects.select_for_update().get(farmacia=sale.farmacia, medicamento=sale.medicamento)
        remaining = quantity
        previous_returns = {}
        for item in DevolucionVentaLote.objects.filter(devolucion__venta=sale).values("lote_id").annotate(total=Sum("cantidad")):
            previous_returns[item["lote_id"]] = item["total"]
        for sold in sale.lotes_consumidos.select_related("lote").order_by("lote__fecha_vencimiento"):
            available = sold.cantidad - previous_returns.get(sold.lote_id, 0)
            restored = min(remaining, available)
            if restored <= 0: continue
            lot = Lote.objects.select_for_update().get(pk=sold.lote_id); lot.cantidad_disponible += restored; lot.save()
            DevolucionVentaLote.objects.create(devolucion=returned_record, lote=lot, cantidad=restored)
            MovimientoInventario.objects.create(farmacia=sale.farmacia, medicamento=sale.medicamento, lote=lot, tipo="ENTRADA", concepto="DEVOLUCION_CLIENTE", cantidad=restored, saldo_anterior=inventory.stock_actual, saldo_nuevo=inventory.stock_actual+restored, costo_unitario=lot.costo_unitario, documento_tipo="DEVOLUCION", documento_id=returned_record.id, usuario=request.user, observacion=reason)
            inventory.stock_actual += restored; remaining -= restored
            if not remaining: break
        inventory.save()
        note = NotaCredito.objects.create(devolucion=returned_record, numero=f"NC-{uuid4().hex[:12].upper()}", motivo=reason, total=refund)
        if sale.sesion_caja and not sale.sesion_caja.cerrada_en:
            MovimientoCaja.objects.create(sesion=sale.sesion_caja, tipo="DEVOLUCION", forma_pago=sale.forma_pago, monto=refund, referencia=note.numero, usuario=request.user, observacion=reason)
        audit(request, AuditLog.Accion.CREAR, "devolucion_cliente", returned_record, f"Devolución de venta {sale.referencia_cliente}", sale.farmacia, {"cantidad": quantity, "nota_credito": note.numero})
        return Response({"id": returned_record.id, "saleId": sale.referencia_cliente, "quantity": quantity, "total": float(refund), "creditNote": note.numero, "stock": inventory.stock_actual}, status=201)


class LotAlertsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today = timezone.localdate()
        days = min(max(int(request.query_params.get("days", 90)), 1), 365)
        rows = Lote.objects.filter(cantidad_disponible__gt=0).select_related("farmacia", "medicamento")
        if not (request.user.is_staff or request.user.is_superuser):
            rows = rows.filter(farmacia__usuarios_asignados__usuario=request.user, farmacia__usuarios_asignados__activo=True)
        data = [{
            "id": lot.id, "branchId": lot.farmacia.codigo.lower(), "sku": lot.medicamento.codigo_interno,
            "product": lot.medicamento.nombre_comercial, "number": lot.numero,
            "expires": lot.fecha_vencimiento.isoformat(), "quantity": lot.cantidad_disponible,
            "status": "EXPIRED" if lot.fecha_vencimiento < today else "EXPIRING",
        } for lot in rows.filter(fecha_vencimiento__lte=today + timedelta(days=days)).order_by("fecha_vencimiento")]
        return Response({"lots": data})


class ExcelReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not (request.user.is_staff or request.user.is_superuser):
            return Response({"detail": "Solo administradores pueden exportar reportes."}, status=403)
        branch = request.query_params.get("branch")
        date_from = request.query_params.get("from")
        date_to = request.query_params.get("to")
        user = request.query_params.get("user")
        movement_type = request.query_params.get("movement")

        inventories = InventarioFarmacia.objects.select_related("farmacia", "medicamento")
        sales = Venta.objects.select_related("farmacia", "medicamento", "usuario").prefetch_related("lotes_consumidos__lote")
        movements = MovimientoInventario.objects.select_related("farmacia", "medicamento", "lote", "usuario")
        if branch and branch != "ALL":
            inventories = inventories.filter(farmacia__codigo__iexact=branch)
            sales = sales.filter(farmacia__codigo__iexact=branch)
            movements = movements.filter(farmacia__codigo__iexact=branch)
        if date_from:
            sales = sales.filter(creado_en__date__gte=date_from); movements = movements.filter(creado_en__date__gte=date_from)
        if date_to:
            sales = sales.filter(creado_en__date__lte=date_to); movements = movements.filter(creado_en__date__lte=date_to)
        if user:
            sales = sales.filter(usuario_id=user); movements = movements.filter(usuario_id=user)
        if movement_type:
            movements = movements.filter(tipo=movement_type)

        workbook = Workbook(); inventory_sheet = workbook.active; inventory_sheet.title = "Inventario"
        self._sheet(inventory_sheet, ["Sucursal", "SKU", "Medicamento", "Stock", "Mínimo", "Compra", "Venta"], [
            [x.farmacia.nombre, x.medicamento.codigo_interno, x.medicamento.nombre_comercial, x.stock_actual, x.stock_minimo, x.precio_compra, x.precio_venta] for x in inventories
        ])
        sale_sheet = workbook.create_sheet("Ventas")
        self._sheet(sale_sheet, ["Fecha", "Sucursal", "Referencia", "Usuario", "Medicamento", "Lotes FEFO", "Cantidad", "Precio", "Total"], [
            [timezone.localtime(x.creado_en).strftime("%Y-%m-%d %H:%M"), x.farmacia.nombre, x.referencia_cliente, x.usuario.email,
             x.medicamento.nombre_comercial, ", ".join(f"{d.lote.numero} ({d.cantidad})" for d in x.lotes_consumidos.all()), x.cantidad, x.precio_unitario, x.total] for x in sales
        ])
        movement_sheet = workbook.create_sheet("Movimientos")
        self._sheet(movement_sheet, ["Fecha", "Sucursal", "Tipo", "Concepto", "Usuario", "Medicamento", "Lote", "Cantidad", "Saldo anterior", "Saldo nuevo"], [
            [timezone.localtime(x.creado_en).strftime("%Y-%m-%d %H:%M"), x.farmacia.nombre, x.tipo, x.concepto, x.usuario.email,
             x.medicamento.nombre_comercial, x.lote.numero if x.lote else "", x.cantidad, x.saldo_anterior, x.saldo_nuevo] for x in movements
        ])
        stream = BytesIO(); workbook.save(stream)
        response = HttpResponse(stream.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="reporte-supabase-{today_string()}.xlsx"'
        return response

    @staticmethod
    def _sheet(sheet, headers, rows):
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="2563EB")
        for row in rows:
            sheet.append(row)
        sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(c.value or "")) for c in column) + 2, 45)


def today_string():
    return timezone.localdate().isoformat()


class DashboardView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        branch = request.query_params.get("branch")
        inventories = InventarioFarmacia.objects.select_related("farmacia", "medicamento")
        sales = Venta.objects.filter(anulada=False)
        lots = Lote.objects.filter(cantidad_disponible__gt=0)
        if branch and branch != "ALL":
            inventories=inventories.filter(farmacia__codigo__iexact=branch); sales=sales.filter(farmacia__codigo__iexact=branch); lots=lots.filter(farmacia__codigo__iexact=branch)
        today=timezone.localdate(); month_start=today.replace(day=1)
        value_expr=ExpressionWrapper(F("stock_actual")*F("precio_compra"),output_field=DecimalField(max_digits=18,decimal_places=2))
        expired_expr=ExpressionWrapper(F("cantidad_disponible")*F("costo_unitario"),output_field=DecimalField(max_digits=18,decimal_places=2))
        branch_rows=[]
        for pharmacy in Farmacia.objects.filter(activo=True):
            inv=inventories.filter(farmacia=pharmacy); branch_sales=sales.filter(farmacia=pharmacy,creado_en__date__gte=month_start)
            branch_rows.append({"id":pharmacy.codigo.lower(),"name":pharmacy.nombre,"stock":sum(x.stock_actual for x in inv),"inventoryValue":float(inv.aggregate(v=Sum(value_expr))["v"] or 0),"sales":float(branch_sales.aggregate(v=Sum("total"))["v"] or 0)})
        return Response({"salesToday":float(sales.filter(creado_en__date=today).aggregate(v=Sum("total"))["v"] or 0),"salesMonth":float(sales.filter(creado_en__date__gte=month_start).aggregate(v=Sum("total"))["v"] or 0),"inventoryValue":float(inventories.aggregate(v=Sum(value_expr))["v"] or 0),"expiredLoss":float(lots.filter(fecha_vencimiento__lt=today).aggregate(v=Sum(expired_expr))["v"] or 0),"critical":inventories.filter(stock_actual__lte=F("stock_minimo")).count(),"branches":branch_rows})
